"""Build a normalized EIA-861 2024 master dataset.

Reads each EIA-861 2024 workbook from the repo root, locates the header row
by scanning for "Data Year" / "Year" in column A, flattens multi-row headers,
writes a cleaned CSV per source sheet to data/eia861_2024/, and builds a
single utility x state wide master table (master_utility_2024.csv) that joins
the most actionable metrics onto the Frame_2024 utility spine.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "data" / "eia861_2024"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Map of source workbook -> sheets to extract. None means all sheets.
SOURCES: dict[str, list[str] | None] = {
    "Frame_2024.xlsx": None,
    "Utility_Data_2024.xlsx": None,
    "Operational_Data_2024.xlsx": None,
    "Short_Form_2024.xlsx": None,
    "Balancing_Authority_2024.xlsx": None,
    "Service_Territory_2024.xlsx": None,
    "Delivery_Companies_2024.xlsx": None,
    "Distribution_Systems_2024.xlsx": None,
    "Reliability_2024.xlsx": None,
    "Sales_Ult_Cust_2024.xlsx": None,
    "Sales_Ult_Cust_CS_2024.xlsx": None,
    "Mergers_2024.xlsx": None,
    "Energy_Efficiency_2024.xlsx": None,
    "Demand_Response_2024.xlsx": None,
    "Dynamic_Pricing_2024.xlsx": None,
    "Advanced_Meters_2024.xlsx": None,
    "Net_Metering_2024.xlsx": None,
    "Non_Net_Metering_Distributed_2024.xlsx": None,
}

KEY_HEADER_TOKENS = {"data year", "year"}


def _slug(name: str) -> str:
    name = re.sub(r"[^0-9A-Za-z]+", "_", str(name)).strip("_").lower()
    return name or "col"


def _find_header_row(ws) -> int:
    """Return 0-indexed row containing 'Data Year' or 'Year' in col A."""
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        cell = row[0]
        if cell is None:
            continue
        if str(cell).strip().lower() in KEY_HEADER_TOKENS:
            return idx
    return 0


def _flatten_headers(header_rows: list[tuple]) -> list[str]:
    """Forward-fill upper header rows and concat with the bottom header."""
    if not header_rows:
        return []
    width = max(len(r) for r in header_rows)
    padded = [list(r) + [None] * (width - len(r)) for r in header_rows]
    # forward fill all but last row
    for r in padded[:-1]:
        last = ""
        for i, v in enumerate(r):
            if v is None or str(v).strip() == "":
                r[i] = last
            else:
                last = str(v).strip()
    bottom = padded[-1]
    out = []
    for i in range(width):
        parts = []
        for r in padded[:-1]:
            v = r[i]
            if v and str(v).strip():
                parts.append(str(v).strip())
        b = bottom[i]
        if b and str(b).strip():
            parts.append(str(b).strip())
        name = " | ".join(dict.fromkeys(parts))  # dedupe keep order
        out.append(name or f"col_{i}")
    # uniquify
    seen: dict[str, int] = {}
    uniq = []
    for n in out:
        s = _slug(n)
        if s in seen:
            seen[s] += 1
            uniq.append(f"{s}_{seen[s]}")
        else:
            seen[s] = 0
            uniq.append(s)
    return uniq


_ID_RENAMES = {
    "utility_characteristics_data_year": "data_year",
    "utility_characteristics_utility_number": "utility_number",
    "utility_characteristics_utility_name": "utility_name",
    "utility_characteristics_state": "state",
    "utility_characteristics_ownership": "ownership",
    "utility_characteristics_ba_code": "ba_code",
    "utility_characteristics_short_form": "short_form",
    "utility_characteristics_part": "part",
    "utility_characteristics_service_type": "service_type",
    "utility_characteristics_data_type_o_observed_i_imputed": "data_type",
    "characteristics_data_year": "data_year",
    "characteristics_state": "state",
}


def _normalize_id_cols(df: pd.DataFrame) -> pd.DataFrame:
    return df.rename(columns=_ID_RENAMES)


def read_sheet(xlsx: Path, sheet: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[sheet]
    hdr_idx = _find_header_row(ws)
    header_rows: list[tuple] = []
    data_rows: list[tuple] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < hdr_idx:
            if i >= max(0, hdr_idx - 2):  # keep up to 2 supra-headers
                header_rows.append(row)
        elif i == hdr_idx:
            header_rows.append(row)
        else:
            data_rows.append(row)
    wb.close()
    cols = _flatten_headers(header_rows)
    width = len(cols)
    norm_rows = [list(r) + [None] * (width - len(r)) if len(r) < width else list(r)[:width] for r in data_rows]
    df = pd.DataFrame(norm_rows, columns=cols)
    # drop fully-empty rows
    df = df.dropna(how="all").reset_index(drop=True)
    # normalize "." sentinel -> NaN
    df = df.replace({".": pd.NA, "": pd.NA})
    df = _normalize_id_cols(df)
    return df


def coerce_numeric(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")


def process_all() -> tuple[dict, dict]:
    summary: dict = {}
    tables: dict[str, dict[str, pd.DataFrame]] = {}
    for fname in SOURCES:
        path = ROOT / fname
        if not path.exists():
            print(f"  ! missing {fname}")
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        stem = path.stem.lower()
        tables[stem] = {}
        per_file = {}
        for sh in sheets:
            df = read_sheet(path, sh)
            out_name = f"{stem}__{_slug(sh)}.csv"
            df.to_csv(OUT_DIR / out_name, index=False)
            tables[stem][_slug(sh)] = df
            per_file[sh] = {
                "rows": int(len(df)),
                "cols": int(len(df.columns)),
                "output": out_name,
                "columns": list(df.columns)[:60],
            }
            print(f"  - {fname} :: {sh}  rows={len(df)} cols={len(df.columns)}")
        summary[fname] = per_file
    return summary, tables


def _pick(df: pd.DataFrame, *needles: str) -> str | None:
    """Return first column whose name contains ALL needles (case-insensitive)."""
    low = {c: c.lower() for c in df.columns}
    for c, lc in low.items():
        if all(n in lc for n in needles):
            return c
    return None


def build_master(tables: dict) -> pd.DataFrame:
    # Spine = Frame (utility identification)
    frame = tables["frame_2024"]["frame"].copy()
    frame = frame.rename(columns={
        "utility_number": "utility_id",
        "utility_name": "utility_name",
        "ownership_code": "ownership_code",
        "ownership": "ownership",
        "data_year": "year",
    })
    keep_cols = ["year", "utility_id", "utility_name", "ownership_code", "ownership", "short_form"]
    frame = frame[[c for c in keep_cols if c in frame.columns]]
    frame["utility_id"] = coerce_numeric(frame["utility_id"]).astype("Int64")
    master = frame.copy()

    # --- Service Territory: count of counties + list of states served
    st = pd.concat([
        tables["service_territory_2024"]["counties_states"],
        tables["service_territory_2024"]["counties_territories"],
    ], ignore_index=True)
    st["utility_id"] = coerce_numeric(st["utility_number"]).astype("Int64")
    agg_st = st.groupby("utility_id").agg(
        n_counties=("county", "count"),
        states_served=("state", lambda s: ",".join(sorted(set(map(str, s.dropna()))))),
    ).reset_index()
    master = master.merge(agg_st, on="utility_id", how="left")

    # --- Sales to Ultimate Customers (states + territories)
    sales = pd.concat([
        tables["sales_ult_cust_2024"]["states"],
        tables["sales_ult_cust_2024"]["territories"],
    ], ignore_index=True)
    sales["utility_id"] = coerce_numeric(sales["utility_number"]).astype("Int64")

    def agg_sales(prefix: str) -> dict[str, pd.Series]:
        out = {}
        for sector in ["residential", "commercial", "industrial", "transportation", "total"]:
            rev_col = _pick(sales, sector, "thousand")
            mwh_col = _pick(sales, sector, "megawatt")
            cust_col = _pick(sales, sector, "customer")
            # The flattened names include sector prefix from supra-headers
            # so try sector-scoped search
            for c in sales.columns:
                lc = c.lower()
                if sector in lc:
                    if "thousand" in lc and f"{prefix}_revenue_{sector}_k$" not in out:
                        out[f"{prefix}_revenue_{sector}_k$"] = coerce_numeric(sales[c])
                    elif ("megawatt" in lc or "mwh" in lc) and f"{prefix}_sales_{sector}_mwh" not in out:
                        out[f"{prefix}_sales_{sector}_mwh"] = coerce_numeric(sales[c])
                    elif "customer" in lc and f"{prefix}_customers_{sector}" not in out:
                        out[f"{prefix}_customers_{sector}"] = coerce_numeric(sales[c])
        return out

    sales_metrics = pd.DataFrame({"utility_id": sales["utility_id"]})
    for k, v in agg_sales("sales").items():
        sales_metrics[k] = v
    sales_agg = sales_metrics.groupby("utility_id").sum(min_count=1).reset_index()
    master = master.merge(sales_agg, on="utility_id", how="left")

    # Total revenue / sales / customers convenience columns
    for metric in ["revenue", "sales", "customers"]:
        cols = [c for c in master.columns if c.startswith(f"sales_{metric}_") and not c.endswith("_total_k$") and not c.endswith("_total_mwh") and not c.endswith("_total")]
        sector_cols = [c for c in cols if not c.endswith("total")]
        if sector_cols:
            master[f"sales_{metric}_total"] = master[sector_cols].sum(axis=1, min_count=1)

    # --- Advanced Meters: AMI + AMR counts
    am = pd.concat([
        tables["advanced_meters_2024"]["states"],
        tables["advanced_meters_2024"]["territories"],
    ], ignore_index=True)
    am["utility_id"] = coerce_numeric(am["utility_number"]).astype("Int64")
    am_metrics = pd.DataFrame({"utility_id": am["utility_id"]})
    for c in am.columns:
        lc = c.lower()
        if "ami" in lc and any(s in lc for s in ["residential", "commercial", "industrial", "transportation"]):
            am_metrics[f"ami_{_slug(c)}"] = coerce_numeric(am[c])
        if "amr" in lc and any(s in lc for s in ["residential", "commercial", "industrial", "transportation"]):
            am_metrics[f"amr_{_slug(c)}"] = coerce_numeric(am[c])
    am_agg = am_metrics.groupby("utility_id").sum(min_count=1).reset_index()
    ami_cols = [c for c in am_agg.columns if c.startswith("ami_")]
    amr_cols = [c for c in am_agg.columns if c.startswith("amr_")]
    am_agg["ami_meters_total"] = am_agg[ami_cols].sum(axis=1, min_count=1) if ami_cols else pd.NA
    am_agg["amr_meters_total"] = am_agg[amr_cols].sum(axis=1, min_count=1) if amr_cols else pd.NA
    master = master.merge(am_agg[["utility_id", "ami_meters_total", "amr_meters_total"]], on="utility_id", how="left")

    # --- Demand Response
    dr = pd.concat([
        tables["demand_response_2024"]["demand_response_states"],
        tables["demand_response_2024"]["demand_response_territories"],
    ], ignore_index=True)
    dr["utility_id"] = coerce_numeric(dr["utility_number"]).astype("Int64")
    total_col = _pick(dr, "total")
    if total_col:
        dr_agg = dr.groupby("utility_id").agg(dr_customers_total=(total_col, lambda s: coerce_numeric(s).sum(min_count=1))).reset_index()
        master = master.merge(dr_agg, on="utility_id", how="left")

    # --- Dynamic Pricing (number of customers enrolled by sector + program)
    dp = pd.concat([
        tables["dynamic_pricing_2024"]["dynamic_pricing_states"],
        tables["dynamic_pricing_2024"]["dynamic_pricing_territories"],
    ], ignore_index=True)
    dp["utility_id"] = coerce_numeric(dp["utility_number"]).astype("Int64")
    # sum any "Total" column if present, else sum residential+commercial+industrial+transportation
    sector_cols = []
    for c in dp.columns:
        lc = c.lower()
        if any(s in lc for s in ["residential", "commercial", "industrial", "transportation"]):
            sector_cols.append(c)
    dp_num = dp[["utility_id"] + sector_cols].copy()
    for c in sector_cols:
        dp_num[c] = coerce_numeric(dp_num[c])
    dp_num["dyn_pricing_customers_total"] = dp_num[sector_cols].sum(axis=1, min_count=1)
    dp_agg = dp_num.groupby("utility_id")["dyn_pricing_customers_total"].sum(min_count=1).reset_index()
    master = master.merge(dp_agg, on="utility_id", how="left")

    # --- Net Metering capacity (PV + Wind + Other) MW
    nm = pd.concat([
        tables["net_metering_2024"]["states"],
        tables["net_metering_2024"]["territories"],
    ], ignore_index=True)
    nm["utility_id"] = coerce_numeric(nm["utility_number"]).astype("Int64")
    cap_cols = [c for c in nm.columns if "capacity" in c.lower() and any(s in c.lower() for s in ["residential", "commercial", "industrial", "transportation"])]
    nm_num = nm[["utility_id"] + cap_cols].copy()
    for c in cap_cols:
        nm_num[c] = coerce_numeric(nm_num[c])
    nm_num["net_metering_capacity_mw"] = nm_num[cap_cols].sum(axis=1, min_count=1)
    nm_agg = nm_num.groupby("utility_id")["net_metering_capacity_mw"].sum(min_count=1).reset_index()
    master = master.merge(nm_agg, on="utility_id", how="left")

    # --- Reliability (SAIDI/SAIFI - take "All Events" / IEEE Standard SAIDI minutes per year)
    rel = pd.concat([
        tables["reliability_2024"]["reliability_states"],
        tables["reliability_2024"]["reliability_territories"],
    ], ignore_index=True)
    rel["utility_id"] = coerce_numeric(rel["utility_number"]).astype("Int64")
    saidi_col = next((c for c in rel.columns if "saidi" in c.lower() and "all_events" in c.lower()), None) \
        or next((c for c in rel.columns if "saidi" in c.lower()), None)
    saifi_col = next((c for c in rel.columns if "saifi" in c.lower() and "all_events" in c.lower()), None) \
        or next((c for c in rel.columns if "saifi" in c.lower()), None)
    if saidi_col and saifi_col:
        rel_agg = rel.groupby("utility_id").agg(
            saidi_minutes_per_year=(saidi_col, lambda s: coerce_numeric(s).mean()),
            saifi_per_year=(saifi_col, lambda s: coerce_numeric(s).mean()),
        ).reset_index()
        master = master.merge(rel_agg, on="utility_id", how="left")

    # --- Operational Data (peak demand, generation)
    op = pd.concat([
        tables["operational_data_2024"]["states"],
        tables["operational_data_2024"]["territories"],
    ], ignore_index=True)
    op["utility_id"] = coerce_numeric(op["utility_number"]).astype("Int64")
    cols = {
        "summer_peak_mw": _pick(op, "summer", "peak"),
        "winter_peak_mw": _pick(op, "winter", "peak"),
        "net_generation_mwh": _pick(op, "net", "generation"),
        "wholesale_purchases_mwh": _pick(op, "wholesale", "purchase"),
    }
    op_keep = ["utility_id"] + [c for c in cols.values() if c]
    op_sub = op[op_keep].copy()
    for c in op_keep[1:]:
        op_sub[c] = coerce_numeric(op_sub[c])
    op_sub = op_sub.groupby("utility_id").sum(min_count=1).reset_index()
    rename = {v: k for k, v in cols.items() if v}
    op_sub = op_sub.rename(columns=rename)
    master = master.merge(op_sub, on="utility_id", how="left")

    # --- Utility Data (NERC region)
    ud = pd.concat([
        tables["utility_data_2024"]["states"],
        tables["utility_data_2024"]["territories"],
    ], ignore_index=True)
    ud["utility_id"] = coerce_numeric(ud["utility_number"]).astype("Int64")
    nerc_col = _pick(ud, "nerc")
    if nerc_col:
        ud_keep = ud.groupby("utility_id").agg(nerc_region=(nerc_col, lambda s: s.dropna().astype(str).iloc[0] if s.notna().any() else None)).reset_index()
        master = master.merge(ud_keep, on="utility_id", how="left")

    # --- Energy Efficiency (total incremental savings)
    ee = pd.concat([
        tables["energy_efficiency_2024"]["energy_efficiency_states"],
        tables["energy_efficiency_2024"]["energy_efficiency_territories"],
    ], ignore_index=True)
    ee["utility_id"] = coerce_numeric(ee["utility_number"]).astype("Int64")
    tot_col = _pick(ee, "total")
    if tot_col:
        ee_agg = ee.groupby("utility_id").agg(ee_incremental_savings_mwh=(tot_col, lambda s: coerce_numeric(s).sum(min_count=1))).reset_index()
        master = master.merge(ee_agg, on="utility_id", how="left")

    # Order columns: identifiers first, then categories
    master = master.sort_values("utility_id").reset_index(drop=True)
    return master


def main():
    print("Reading source workbooks ...")
    summary, tables = process_all()

    print("\nBuilding master utility table ...")
    master = build_master(tables)
    master_path = OUT_DIR / "master_utility_2024.csv"
    master.to_csv(master_path, index=False)
    print(f"  -> {master_path}  rows={len(master)} cols={len(master.columns)}")

    # Write a summary JSON
    (OUT_DIR / "_source_summary.json").write_text(json.dumps(summary, indent=2, default=str))

    # Console summary of master
    print("\nMaster columns:")
    for c in master.columns:
        print(f"  - {c}")
    print(f"\nUtilities: {master['utility_id'].nunique()}")


if __name__ == "__main__":
    main()
